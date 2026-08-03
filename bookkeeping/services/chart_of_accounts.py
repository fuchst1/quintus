import hashlib
import json
from datetime import datetime
from io import BytesIO
from pathlib import Path
from typing import Any

from django.core.files.base import ContentFile
from django.db import transaction

from bookkeeping.models import KontenplanEintrag, KontenplanVersion, Mandant
from bookkeeping.services.audit import record_audit_event


class ReferenceValidationError(ValueError):
    """Raised when a local bookkeeping reference file violates its contract."""


EXPECTED_WORKSHEETS = ("Allgemeines", "Eingaben", "Auswertung", "UVA", "CSV", "Kontenplan")
REQUIRED_TRANSACTION_FIELDS = ("booking", "valuation", "partnerName", "partnerAccount", "referenceNumber", "amount")


def _load_openpyxl():
    try:
        import openpyxl
    except ImportError as exc:  # pragma: no cover - handled by deployment dependency
        raise ReferenceValidationError("openpyxl ist nicht installiert.") from exc
    return openpyxl


def _read_json_bytes(payload: bytes) -> list[dict[str, Any]]:
    try:
        data = json.loads(payload.decode("utf-8"))
    except (UnicodeDecodeError, json.JSONDecodeError) as exc:
        raise ReferenceValidationError("Die Bankdatei ist kein gültiges UTF-8-JSON.") from exc
    if not isinstance(data, list):
        raise ReferenceValidationError("Die Bankdatei muss eine Liste von Transaktionen enthalten.")
    return data


def validate_monthly_bank_json(payload: bytes, *, expected_transactions: int | None = None) -> dict[str, Any]:
    """Validate a compatible monthly bank JSON file without importing it.

    The transaction count is reported, not fixed. An expected count is only an
    explicit local acceptance assertion.
    """
    transactions = _read_json_bytes(payload)
    references: set[str] = set()
    precisions: set[int] = set()
    currencies: set[str] = set()

    for position, transaction in enumerate(transactions, start=1):
        if not isinstance(transaction, dict):
            raise ReferenceValidationError(f"Transaktion {position} ist kein Objekt.")
        missing = [field for field in REQUIRED_TRANSACTION_FIELDS if field not in transaction]
        if missing:
            raise ReferenceValidationError(f"Transaktion {position} enthält nicht: {', '.join(missing)}.")

        reference = transaction["referenceNumber"]
        if not isinstance(reference, str) or not reference.strip():
            raise ReferenceValidationError(f"Transaktion {position} hat keine gültige referenceNumber.")
        if reference in references:
            raise ReferenceValidationError(f"referenceNumber ist nicht eindeutig: {reference}.")
        references.add(reference)

        for field in ("booking", "valuation"):
            value = transaction[field]
            if not isinstance(value, str):
                raise ReferenceValidationError(f"Transaktion {position}: {field} muss ein ISO-Datum sein.")
            try:
                datetime.fromisoformat(value.replace("Z", "+00:00"))
            except ValueError as exc:
                raise ReferenceValidationError(
                    f"Transaktion {position}: {field} ist kein ISO-Datum mit Zeitzone."
                ) from exc

        if not isinstance(transaction["partnerName"], str) or not transaction["partnerName"].strip():
            raise ReferenceValidationError(f"Transaktion {position} hat keinen Partnernamen.")
        partner_account = transaction["partnerAccount"]
        if (
            not isinstance(partner_account, dict)
            or not isinstance(partner_account.get("iban"), str)
            or not partner_account["iban"].strip()
        ):
            raise ReferenceValidationError(f"Transaktion {position} hat keine Partner-IBAN.")

        amount = transaction["amount"]
        if not isinstance(amount, dict) or not isinstance(amount.get("value"), int):
            raise ReferenceValidationError(f"Transaktion {position} hat keinen ganzzahligen amount.value.")
        precision = amount.get("precision")
        if not isinstance(precision, int) or precision < 0 or precision > 6:
            raise ReferenceValidationError(f"Transaktion {position} hat eine ungültige Betragspräzision.")
        currency = amount.get("currency")
        if currency != "EUR":
            raise ReferenceValidationError(f"Transaktion {position} hat eine nicht unterstützte Währung: {currency!r}.")
        precisions.add(precision)
        currencies.add(currency)

    count = len(transactions)
    if expected_transactions is not None and count != expected_transactions:
        raise ReferenceValidationError(
            f"Erwartet wurden {expected_transactions} Transaktionen, erkannt wurden {count}."
        )
    return {
        "transaction_count": count,
        "currencies": sorted(currencies),
        "amount_precisions": sorted(precisions),
        "sha256": hashlib.sha256(payload).hexdigest(),
    }


def validate_chart_of_accounts_workbook(payload: bytes) -> dict[str, Any]:
    """Read the immutable workbook contract and extract Kontenplan column D."""
    openpyxl = _load_openpyxl()
    try:
        workbook = openpyxl.load_workbook(BytesIO(payload), read_only=True, data_only=False)
    except Exception as exc:  # openpyxl exposes several implementation exceptions
        raise ReferenceValidationError("Die Vorlage kann nicht als Excel-Arbeitsmappe gelesen werden.") from exc

    missing_sheets = [sheet for sheet in EXPECTED_WORKSHEETS if sheet not in workbook.sheetnames]
    if missing_sheets:
        raise ReferenceValidationError(f"Die Vorlage enthält nicht: {', '.join(missing_sheets)}.")
    inputs = workbook["Eingaben"]
    if inputs.max_column < 19 or inputs.max_row < 8:
        raise ReferenceValidationError("Das Sheet Eingaben besitzt nicht die erwartete Struktur.")

    accounts = workbook["Kontenplan"]
    entries: list[dict[str, str]] = []
    seen_categories: set[str] = set()
    for row in accounts.iter_rows(min_row=2, values_only=True):
        category = row[3] if len(row) > 3 else None
        if category is None or not str(category).strip():
            continue
        category_text = str(category).strip()
        if category_text in seen_categories:
            raise ReferenceValidationError(f"Kategorie ist im Kontenplan doppelt: {category_text}.")
        seen_categories.add(category_text)
        values = ["" if value is None else str(value).strip() for value in row]
        entries.append(
            {
                "kategorie_text": category_text,
                "kontonummer": values[0] if len(values) > 0 else "",
                "bezeichnung": values[1] if len(values) > 1 else "",
                "kontoart": values[2] if len(values) > 2 else "",
                "kontoklasse": values[4] if len(values) > 4 else "",
                "ust_stcode": values[5] if len(values) > 5 else "",
            }
        )
    if not entries:
        raise ReferenceValidationError("Der Kontenplan enthält keine Kategorien in Spalte D.")
    return {
        "sheet_names": list(workbook.sheetnames),
        "input_max_row": inputs.max_row,
        "input_max_column": inputs.max_column,
        "entry_count": len(entries),
        "entries": entries,
        "sha256": hashlib.sha256(payload).hexdigest(),
    }


def import_chart_of_accounts(
    *,
    mandant: Mandant,
    bezeichnung: str,
    gueltig_ab,
    uploaded_file,
    user: Any = None,
) -> KontenplanVersion:
    """Persist an immutable original workbook and its extracted chart atomically."""
    payload = uploaded_file.read()
    uploaded_file.seek(0)
    validation = validate_chart_of_accounts_workbook(payload)
    original_filename = Path(uploaded_file.name).name
    if KontenplanVersion.objects.filter(mandant=mandant, vorlage_sha256=validation["sha256"]).exists():
        raise ReferenceValidationError("Diese unveränderte Kontenplanvorlage wurde bereits importiert.")
    version: KontenplanVersion | None = None
    stored_name = ""
    try:
        with transaction.atomic():
            KontenplanVersion.objects.filter(mandant=mandant, aktiv=True).update(aktiv=False)
            version = KontenplanVersion(
                mandant=mandant,
                bezeichnung=bezeichnung,
                gueltig_ab=gueltig_ab,
                vorlage_dateiname=original_filename,
                vorlage_sha256=validation["sha256"],
                aktiv=True,
            )
            version.vorlage_datei.save(original_filename, ContentFile(payload), save=False)
            stored_name = version.vorlage_datei.name
            version.save()
            KontenplanEintrag.objects.bulk_create(
                [KontenplanEintrag(version=version, **entry) for entry in validation["entries"]],
                batch_size=500,
            )
            record_audit_event(
                mandant=mandant,
                objekt_typ="KontenplanVersion",
                objekt_id=version.pk,
                aktion="kontenplan_importiert",
                nachher={"sha256": validation["sha256"], "eintraege": validation["entry_count"]},
                user=user,
            )
    except Exception:
        if stored_name and version is not None:
            version.vorlage_datei.storage.delete(stored_name)
        raise
    return version
