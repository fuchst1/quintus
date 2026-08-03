from io import BytesIO

from openpyxl import Workbook


def make_workbook_bytes(*, categories=("Erlöse", "Büro"), include_accounts=True) -> bytes:
    workbook = Workbook()
    general = workbook.active
    general.title = "Allgemeines"
    inputs = workbook.create_sheet("Eingaben")
    workbook.create_sheet("Auswertung")
    workbook.create_sheet("UVA")
    workbook.create_sheet("CSV")
    accounts = workbook.create_sheet("Kontenplan")
    for column in range(1, 20):
        inputs.cell(7, column, f"Spalte {column}")
    inputs.cell(8, 9, "=G8*1")
    inputs.cell(8, 10, "=G8*2")
    inputs.cell(8, 12, "=G8*3")
    if include_accounts:
        accounts.append(["Kontonummer", "Bezeichnung", "Kontoart", "Kategorie", "Kontoklasse", "USt"])
        for index, category in enumerate(categories, start=1):
            accounts.append([f"{index:04d}", f"Konto {index}", "Erfolg", category, "Klasse", "20"])
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def make_bank_json(*, count=2) -> bytes:
    import json

    return json.dumps(
        [
            {
                "booking": "2026-07-01T00:00:00.000+0200",
                "valuation": "2026-07-01T00:00:00.000+0200",
                "partnerName": f"Partner {index}",
                "partnerAccount": {"iban": f"AT61190430023457320{index:02d}"},
                "referenceNumber": f"REF-{index}",
                "amount": {"value": index * 100, "precision": 2, "currency": "EUR"},
            }
            for index in range(1, count + 1)
        ]
    ).encode("utf-8")
